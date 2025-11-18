import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import csv
import sqlite3
import os
import sys
from tkcalendar import DateEntry 
# Importação da biblioteca openpyxl para ler e ESCREVER arquivos Excel
from openpyxl import load_workbook, Workbook 

# ----------------------------------------------------
# --- VARIÁVEIS GLOBAIS DE ESTADO E DATA ---
# ----------------------------------------------------

# Variáveis para armazenar o último filtro aplicado (necessário para a exportação)
LAST_FILTER_START_DATE_BR = ""
LAST_FILTER_END_DATE_BR = ""

# ----------------------------------------------------
# --- FUNÇÕES AUXILIARES DE DATA E VALOR ---
# ----------------------------------------------------

def to_iso(date_obj_br):
    """Converte objeto datetime ou string DD/MM/AAAA para YYYY-MM-DD para o SQLite."""
    if not date_obj_br:
        return None
    
    if isinstance(date_obj_br, datetime):
        # Se for um objeto datetime (vindo do DateEntry ou Excel)
        return date_obj_br.strftime('%Y-%m-%d')
    
    date_str_br = str(date_obj_br).strip()
    try:
        # Tenta conversão direta da string DD/MM/AAAA
        return datetime.strptime(date_str_br, '%d/%m/%Y').strftime('%Y-%m-%d')
    except ValueError:
        return None 

def to_br(date_str_iso):
    """Converte YYYY-MM-DD (do SQLite) para DD/MM/AAAA para exibição."""
    if not date_str_iso:
        return ""
    try:
        return datetime.strptime(date_str_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        return date_str_iso 

def clean_and_convert_value(value):
    """Limpa e converte o valor para float, lidando com None/vazio/strings."""
    if value is None or str(value).strip() == '':
        return 0.00
    
    val_str = str(value).strip()
    
    # Remove R$, pontos (milhar) e substitui vírgula (decimal) por ponto
    val_str = val_str.replace('R$', '').replace('.', '').replace(',', '.')
    
    try:
        return float(val_str)
    except ValueError:
        return 0.00 

# ----------------------------------------------------
# --- FUNÇÕES DE BANCO DE DADOS (SQLite) ---
# ----------------------------------------------------

def obter_caminho_db():
    """
    Retorna o caminho correto para o banco de dados.
    """
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), 'restaurantes_atendimentos.db')
    else:
        return 'restaurantes_atendimentos.db'

DATABASE_NAME = obter_caminho_db()

def setup_db():
    """Cria as tabelas de Clientes e Atendimentos se não existirem."""
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cnpj TEXT,
            telefone TEXT,
            endereco TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS atendimentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER NOT NULL,
            data_atendimento DATE NOT NULL,
            descricao TEXT NOT NULL,
            valor REAL,
            FOREIGN KEY (cliente_id) REFERENCES clientes (id)
        )
    ''')
    conn.commit()
    conn.close()

# --- Funções de Clientes (CRUD e Importação) ---
# ... (Funções de Clientes mantidas as originais) ...
def inserir_cliente(nome, cnpj, telefone, endereco):
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO clientes (nome, cnpj, telefone, endereco) VALUES (?, ?, ?, ?)",
        (nome, cnpj, telefone, endereco)
    )
    conn.commit()
    conn.close()

def listar_clientes():
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(clientes)")
    colunas = [info[1] for info in cursor.fetchall()]
    cursor.execute("SELECT id, nome, cnpj, telefone, endereco FROM clientes ORDER BY nome")
    clientes = cursor.fetchall()
    conn.close()
    return colunas, clientes

def editar_cliente(id_cliente, nome, cnpj, telefone, endereco):
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE clientes SET nome=?, cnpj=?, telefone=?, endereco=? WHERE id=?",
        (nome, cnpj, telefone, endereco, id_cliente)
    )
    conn.commit()
    conn.close()

def remover_cliente(id_cliente):
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM atendimentos WHERE cliente_id=?", (id_cliente,))
    cursor.execute("DELETE FROM clientes WHERE id=?", (id_cliente,))
    conn.commit()
    conn.close()

def importar_clientes_excel(filepath):
    """Lê o arquivo Excel e insere os clientes no banco de dados."""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return 0, "A planilha está vazia ou o cabeçalho não foi pulado corretamente."

        clientes_importados = 0
        conn = sqlite3.connect(DATABASE_NAME)
        cursor = conn.cursor()

        for row in rows:
            nome = str(row[0]).strip() if row[0] is not None else None
            cnpj = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
            telefone = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            endereco = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None

            if nome: 
                 cursor.execute(
                    "INSERT INTO clientes (nome, cnpj, telefone, endereco) VALUES (?, ?, ?, ?)",
                    (nome, cnpj, telefone, endereco)
                )
                 clientes_importados += 1

        conn.commit()
        conn.close()
        return clientes_importados, f"{clientes_importados} cliente(s) importado(s) com sucesso!"

    except Exception as e:
        return 0, f"Erro ao ler ou importar o arquivo Excel de Clientes: {e}"


# --- Funções de Atendimentos (CRUD e Importação) ---

def inserir_atendimento(cliente_id, data_atendimento_iso, descricao, valor):
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO atendimentos (cliente_id, data_atendimento, descricao, valor) VALUES (?, ?, ?, ?)",
        (cliente_id, data_atendimento_iso, descricao, valor)
    )
    conn.commit()
    conn.close()

def editar_atendimento(id_atendimento, cliente_id, data_atendimento_iso, descricao, valor):
    """Função para editar um atendimento existente."""
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE atendimentos SET cliente_id=?, data_atendimento=?, descricao=?, valor=? WHERE id=?",
        (cliente_id, data_atendimento_iso, descricao, valor, id_atendimento)
    )
    conn.commit()
    conn.close()

def remover_atendimento(id_atendimento):
    """Função para remover um atendimento."""
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM atendimentos WHERE id=?", (id_atendimento,))
    conn.commit()
    conn.close()

def listar_atendimentos(cliente_id=None, data_inicio_iso=None, data_fim_iso=None):
    """
    Lista atendimentos, filtrando por cliente e/ou por período de datas (ISO YYYY-MM-DD).
    """
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()

    sql = """
        SELECT
            A.id,
            C.nome AS Cliente,
            A.data_atendimento,
            A.descricao,
            A.valor,
            A.cliente_id
        FROM atendimentos AS A
        JOIN clientes AS C ON A.cliente_id = C.id
    """
    params = []
    
    where_clauses = []
    if cliente_id and cliente_id != 'Todos':
        where_clauses.append("A.cliente_id = ?")
        params.append(cliente_id)
    
    if data_inicio_iso:
        where_clauses.append("A.data_atendimento >= ?")
        params.append(data_inicio_iso)
    
    if data_fim_iso:
        where_clauses.append("A.data_atendimento <= ?")
        params.append(data_fim_iso)

    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    
    sql += " ORDER BY A.data_atendimento DESC, A.id DESC"

    cursor.execute(sql, params)
    atendimentos = cursor.fetchall()

    colunas = ['ID', 'Cliente', 'Data', 'Descrição', 'Valor', 'Cliente_ID']
    conn.close()
    return colunas, atendimentos


def importar_atendimentos_excel(filepath, clientes_map):
    """Lê o arquivo Excel de atendimentos e insere no banco de dados."""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return 0, 0, "A planilha está vazia ou o cabeçalho não foi pulado corretamente."

        importados = 0
        erros = 0
        conn = sqlite3.connect(DATABASE_NAME)
        cursor = conn.cursor()
        erros_list = []

        for i, row in enumerate(rows):
            # 0: Data, 1: Cliente, 2: Descrição, 3: Valor (Opcional)
            data_excel = row[0] if len(row) > 0 and row[0] is not None else None
            cliente_nome = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
            descricao = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            valor_excel = row[3] if len(row) > 3 and row[3] is not None else None
            
            linha_num = i + 2 

            cliente_id = clientes_map.get(cliente_nome)
            data_iso = to_iso(data_excel)
            valor_final = clean_and_convert_value(valor_excel)
            
            if not cliente_id:
                erros += 1
                erros_list.append(f"Linha {linha_num}: Cliente '{cliente_nome}' não encontrado no cadastro.")
                continue
            
            if not data_iso:
                erros += 1
                erros_list.append(f"Linha {linha_num}: Data inválida ('{data_excel}'). Esperado DD/MM/AAAA.")
                continue

            if not descricao:
                erros += 1
                erros_list.append(f"Linha {linha_num}: Descrição do atendimento está vazia.")
                continue

            try:
                 cursor.execute(
                    "INSERT INTO atendimentos (cliente_id, data_atendimento, descricao, valor) VALUES (?, ?, ?, ?)",
                    (cliente_id, data_iso, descricao, valor_final)
                )
                 importados += 1
            except Exception as e:
                erros += 1
                erros_list.append(f"Linha {linha_num}: Erro ao inserir no DB: {e}")


        conn.commit()
        conn.close()
        
        msg_sucesso = f"{importados} atendimento(s) importado(s) com sucesso!"
        msg_erros = "\n".join(erros_list) if erros_list else ""
        
        return importados, erros, msg_sucesso + ("\n\n--- Erros ---\n" + msg_erros if erros > 0 else "")

    except Exception as e:
        return 0, 0, f"Erro crítico ao ler ou importar o arquivo Excel de Atendimentos: {e}"


# ----------------------------------------------------
# --- FUNÇÕES DE EXPORTAÇÃO (EXCEL) ---
# ----------------------------------------------------

def formatar_dados_exportacao(tabela):
    """Extrai e formata os dados da Treeview para exportação."""
    data = []
    for child in tabela.get_children():
        row = list(tabela.item(child)["values"])
        # A Treeview contém ['ID', 'Cliente', 'Data', 'Descrição', 'Valor', 'Cliente_ID'] - pegamos os 5 primeiros.
        # Formatamos a data [2] e o valor [4] antes de exportar.
        data.append({
            'DATA': row[2],
            'CLIENTE': row[1],
            'DESCRICAO': row[3],
            'VALOR': row[4].replace('R$ ', '').replace(',', '.')
        })
    return data

def exportar_para_excel(data, data_inicio_br, data_fim_br):
    
    # 1. Definir o título do arquivo
    data_inicio_txt = data_inicio_br.replace('/', '-') if data_inicio_br else "INICIO"
    data_fim_txt = data_fim_br.replace('/', '-') if data_fim_br else "FIM"
    file_name = f"Exportacao_relatorio_Marques_data_{data_inicio_txt}_a_{data_fim_txt}"

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=file_name
    )
    
    if file_path:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Atendimentos"

            # 2. Título do Relatório (Combinado de Células)
            titulo = f"Relatório de Atendimentos Marques - Período: {data_inicio_br if data_inicio_br else 'Início'} a {data_fim_br if data_fim_br else 'Fim'}"
            ws.merge_cells('A1:C1')
            ws['A1'] = titulo
            from openpyxl.styles import Font
            ws['A1'].font = Font(bold=True)
            
            ws.append([]) # Linha vazia
            
            # 3. Cabeçalho das Colunas
            cabecalho = ["DATA", "CLIENTE", "DESCRIÇÃO"]
            ws.append(cabecalho)

            # 4. Inserir os Dados
            for item in data:
                # Filtrar para incluir apenas as colunas solicitadas
                ws.append([
                    item['DATA'], 
                    item['CLIENTE'], 
                    item['DESCRICAO']
                ])

            # Ajuste de largura das colunas (opcional, mas melhora a visualização)
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 80
            
            wb.save(file_path)

            messagebox.showinfo("Sucesso", f"Dados exportados para Excel:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erro de Exportação", f"Não foi possível exportar para Excel:\n{e}")

# ----------------------------------------------------
# --- INÍCIO DA INTERFACE GRÁFICA (Tkinter) ---
# ----------------------------------------------------

setup_db()
root = tk.Tk()
root.title("Sistema de Gestão de Atendimentos TI - Restaurantes")
root.geometry("1100x700")

# --- Configuração do Tema Cinza Escuro ---
BG_COLOR = '#3a3a3a' 
ACCENT_COLOR = '#505050' 
TEXT_COLOR = 'white' 
BUTTON_COLOR = '#606060' 

root.configure(bg=BG_COLOR)

style = ttk.Style()
style.theme_use('clam') 
style.configure(".", background=BG_COLOR, foreground=TEXT_COLOR) 
style.configure("TFrame", background=BG_COLOR)
style.configure("TLabelframe", background=BG_COLOR, foreground=TEXT_COLOR)
style.configure("TLabelframe.Label", background=BG_COLOR, foreground=TEXT_COLOR) 

# Notebook
style.configure("TNotebook", background=BG_COLOR, borderwidth=0)
style.configure("TNotebook.Tab", background=ACCENT_COLOR, foreground=TEXT_COLOR,
                lightcolor=BG_COLOR, darkcolor=BG_COLOR) 
style.map("TNotebook.Tab", background=[("selected", BG_COLOR)],
                           foreground=[("selected", TEXT_COLOR)])

# Combobox
style.configure("TCombobox", fieldbackground=ACCENT_COLOR, background=BUTTON_COLOR,
                foreground=TEXT_COLOR, selectbackground=ACCENT_COLOR, selectforeground=TEXT_COLOR)
style.map("TCombobox", fieldbackground=[("readonly", ACCENT_COLOR)],
                       selectbackground=[("readonly", ACCENT_COLOR)])

# Treeview
style.configure("Treeview", background=ACCENT_COLOR, foreground=TEXT_COLOR,
                fieldbackground=ACCENT_COLOR, rowheight=25)
style.map("Treeview", background=[("selected", BUTTON_COLOR)],
                      foreground=[("selected", "yellow")]) 

style.configure("Treeview.Heading", font=('Arial', 10, 'bold'), background=BUTTON_COLOR,
                foreground=TEXT_COLOR, relief="flat")
style.map("Treeview.Heading", background=[('active', '#707070')]) 

# Scrollbar (para a Treeview)
style.configure("Vertical.TScrollbar", background=ACCENT_COLOR, troughcolor=BG_COLOR,
                bordercolor=ACCENT_COLOR, arrowcolor=TEXT_COLOR)
style.map("Vertical.TScrollbar", background=[('active', BUTTON_COLOR)])

# Buttons
style.configure("TButton", background=BUTTON_COLOR, foreground=TEXT_COLOR,
                font=('Arial', 10), borderwidth=1, relief="raised")
style.map("TButton", background=[('active', '#707070')]) 

# Configurações de Estilos Customizados
style.configure('Danger.TButton', foreground='red', background=BUTTON_COLOR)
style.map('Danger.TButton', background=[('active', '#707070')])
style.configure('Accent.TButton', background='#007acc', foreground='white') # Azul para Importar/Exportar
style.map('Accent.TButton', background=[('active', '#005f99')])

# --- Fim da Configuração do Tema ---

notebook = ttk.Notebook(root)
notebook.pack(pady=10, padx=10, expand=True, fill="both")

CLIENTES_MAP = {}
CLIENTES_ID_MAP = {}


# ------------------------------
# --- ABA 1: ATENDIMENTOS (PRIMEIRA ABA) ---
# ------------------------------
frame_atendimentos = ttk.Frame(notebook)
notebook.add(frame_atendimentos, text="📞 Registro de Atendimentos")


def criar_aba_atendimentos(frame):
    
    # --- Form de Novo Atendimento ---
    form_frame = tk.LabelFrame(frame, text=" Novo Atendimento ", padx=10, pady=10, bg=BG_COLOR, fg=TEXT_COLOR)
    form_frame.pack(fill='x', padx=10, pady=5)
    
    atendimento_id_edit = tk.StringVar(form_frame)
    
    tk.Label(form_frame, text="ID Atendimento (apenas edição):", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=0, column=0, padx=5, pady=5, sticky='w')
    id_entry = tk.Entry(form_frame, width=5, state='readonly', textvariable=atendimento_id_edit,
                        bg=ACCENT_COLOR, fg=TEXT_COLOR, disabledbackground=BG_COLOR, disabledforeground=TEXT_COLOR)
    id_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')

    tk.Label(form_frame, text="Cliente:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=1, column=0, padx=5, pady=5, sticky='w')
    cliente_nomes = list(CLIENTES_MAP.keys())
    cliente_var = tk.StringVar(form_frame)
    cliente_combo = ttk.Combobox(form_frame, textvariable=cliente_var, values=cliente_nomes, state='readonly', width=30)
    if cliente_nomes:
        cliente_combo.set(cliente_nomes[0])
    cliente_combo.grid(row=1, column=1, padx=5, pady=5, sticky='w')
    
    # Substituído tk.Entry por DateEntry
    tk.Label(form_frame, text="Data:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=1, column=2, padx=5, pady=5, sticky='w')
    data_entry = DateEntry(form_frame, width=15, background='darkblue',
                           foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                           locale='pt_BR', selectbackground=BUTTON_COLOR, selectforeground=TEXT_COLOR,
                           normalbackground=ACCENT_COLOR, normalforeground=TEXT_COLOR, headersbackground=BUTTON_COLOR)
    data_entry.grid(row=1, column=3, padx=5, pady=5, sticky='w')
    
    tk.Label(form_frame, text="Valor (R$):", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=2, column=0, padx=5, pady=5, sticky='w')
    valor_entry = tk.Entry(form_frame, width=10, bg=ACCENT_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR)
    valor_entry.insert(0, "0.00")
    valor_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')

    tk.Label(form_frame, text="Descrição do Serviço:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=3, column=0, padx=5, pady=5, sticky='w')
    descricao_text = tk.Text(form_frame, height=3, width=70, bg=ACCENT_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR)
    descricao_text.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky='ew')

    # --- Funções CRUD de Atendimento e Exportação ---

    def limpar_form_atendimento():
        atendimento_id_edit.set("")
        cliente_nomes = list(CLIENTES_MAP.keys())
        cliente_combo.config(values=cliente_nomes)
        cliente_var.set(cliente_nomes[0] if cliente_nomes else "")
        
        # Reseta o DateEntry para a data atual
        data_entry.set_date(datetime.now().date())
        
        valor_entry.delete(0, tk.END)
        valor_entry.insert(0, "0.00")
        descricao_text.delete("1.0", tk.END)
        btn_salvar.config(text="Registrar Novo Atendimento")

    def preencher_form_atendimento(event):
        try:
            limpar_form_atendimento()
            item_selecionado = tabela_atendimentos.selection()[0]
            valores = tabela_atendimentos.item(item_selecionado, 'values')
            
            id_atendimento = valores[0]
            cliente_nome = valores[1]
            data_br = valores[2]
            descricao = valores[3]
            valor_rs = valores[4].replace('R$ ', '').replace(',', '.') 
            
            atendimento_id_edit.set(id_atendimento)
            cliente_var.set(cliente_nome)
            
            # Preenche o DateEntry a partir da data em formato BR (DD/MM/AAAA)
            data_dt = datetime.strptime(data_br, '%d/%m/%Y').date()
            data_entry.set_date(data_dt)

            valor_entry.delete(0, tk.END)
            valor_entry.insert(0, valor_rs)
            descricao_text.delete("1.0", tk.END)
            descricao_text.insert("1.0", descricao)
            
            btn_salvar.config(text=f"Salvar Edição (ID: {id_atendimento})")
        except IndexError:
            pass 
        except ValueError:
            # Lidar com erro se a data do DB estiver inválida
            pass

    def handle_salvar_atendimento():
        id_atendimento = atendimento_id_edit.get()
        cliente_nome = cliente_var.get()
        descricao = descricao_text.get("1.0", tk.END).strip()
        valor_str = valor_entry.get().replace(',', '.')
        
        # Obtém a data como objeto datetime.date
        data_atendimento_dt = data_entry.get_date()
        # Converte para ISO (YYYY-MM-DD) para o DB
        data_iso = to_iso(datetime.combine(data_atendimento_dt, datetime.min.time()))

        valor = clean_and_convert_value(valor_str)

        if not cliente_nome or not descricao or not data_iso:
            messagebox.showerror("Erro", "Cliente, Data e Descrição são obrigatórios!")
            return
        
        cliente_id = CLIENTES_MAP.get(cliente_nome)

        try:
            if id_atendimento:
                editar_atendimento(id_atendimento, cliente_id, data_iso, descricao, valor)
                messagebox.showinfo("Sucesso", f"Atendimento ID {id_atendimento} atualizado!")
            else:
                inserir_atendimento(cliente_id, data_iso, descricao, valor)
                messagebox.showinfo("Sucesso", f"Atendimento para {cliente_nome} registrado!")
            
            limpar_form_atendimento()
            aplicar_filtro() 
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar atendimento:\n{e}")

    def handle_remover_atendimento():
        try:
            item_selecionado = tabela_atendimentos.selection()[0]
            id_atendimento = tabela_atendimentos.item(item_selecionado, 'values')[0]
            
            if messagebox.askyesno("Confirmar Remoção", f"Tem certeza que deseja remover o atendimento ID: {id_atendimento}?"):
                remover_atendimento(id_atendimento)
                messagebox.showinfo("Sucesso", f"Atendimento {id_atendimento} removido.")
                limpar_form_atendimento()
                aplicar_filtro()
        except IndexError:
            messagebox.showwarning("Atenção", "Selecione um atendimento para remover.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível remover o atendimento:\n{e}")
            
    def handle_importar_atendimentos_excel():
        filepath = filedialog.askopenfilename(
            title="Selecione o arquivo Excel de Atendimentos",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if filepath:
            top = tk.Toplevel(root)
            top.title("Importando...")
            tk.Label(top, text="Importação em andamento, por favor aguarde...", bg=BG_COLOR, fg=TEXT_COLOR).pack(padx=20, pady=20)
            root.update_idletasks() 
            
            importados, erros, mensagem = importar_atendimentos_excel(filepath, CLIENTES_MAP)
            top.destroy()
            
            if importados > 0:
                messagebox.showinfo("Importação Concluída", mensagem)
                aplicar_filtro()
            elif erros > 0:
                messagebox.showerror("Importação com Erros", f"Nenhum atendimento importado devido a erros de formatação ou cliente não cadastrado.\n\nDetalhes:\n{mensagem}")
            else:
                 messagebox.showwarning("Importação Vazia", "Nenhum dado válido encontrado para importação.")
    
    def handle_exportar_excel_atendimentos():
        data_exportar = formatar_dados_exportacao(tabela_atendimentos)
        exportar_para_excel(data_exportar, LAST_FILTER_START_DATE_BR, LAST_FILTER_END_DATE_BR)


    # Botões de Ação (Ajustado para incluir o Exportar ao lado do Importar)
    btn_frame = tk.Frame(form_frame, bg=BG_COLOR)
    btn_frame.grid(row=4, column=0, columnspan=4, pady=10)
    
    btn_salvar = ttk.Button(btn_frame, text="Registrar Novo Atendimento", command=handle_salvar_atendimento)
    btn_salvar.pack(side=tk.LEFT, padx=5)
    
    ttk.Button(btn_frame, text="Limpar Formulário", command=limpar_form_atendimento).pack(side=tk.LEFT, padx=5)
    ttk.Button(btn_frame, text="Excluir Atendimento Selecionado", command=handle_remover_atendimento, style='Danger.TButton').pack(side=tk.LEFT, padx=5)
    
    ttk.Button(btn_frame, text="📥 Importar Atendimentos (Excel)", command=handle_importar_atendimentos_excel, style='Accent.TButton').pack(side=tk.LEFT, padx=(20, 5))
    ttk.Button(btn_frame, text="📄 Exportar Relatório (Excel)", command=handle_exportar_excel_atendimentos, style='Accent.TButton').pack(side=tk.LEFT, padx=(5, 5))


    # --- Área de Visualização e Filtros ---
    visual_frame = tk.LabelFrame(frame, text=" Histórico de Atendimentos ", padx=10, pady=10, bg=BG_COLOR, fg=TEXT_COLOR)
    visual_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Filtros
    filter_frame = tk.Frame(visual_frame, bg=BG_COLOR)
    filter_frame.pack(fill='x', pady=5)
    
    tk.Label(filter_frame, text="Filtrar por Cliente:", bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=5)
    filter_cliente_var = tk.StringVar(filter_frame)
    filter_cliente_nomes = ["Todos"] + list(CLIENTES_MAP.keys())
    filter_cliente_combo = ttk.Combobox(filter_frame, textvariable=filter_cliente_var, values=filter_cliente_nomes, state='readonly')
    filter_cliente_combo.set("Todos")
    filter_cliente_combo.pack(side=tk.LEFT, padx=5)
    
    # NOVO FILTRO DE DATA POR PERÍODO (DateEntry)
    tk.Label(filter_frame, text="Data Início:", bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=15)
    filter_data_inicio_entry = DateEntry(filter_frame, width=12, background='darkblue',
                                       foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                                       locale='pt_BR', showweeknumbers=False, selectbackground=BUTTON_COLOR,
                                       selectforeground=TEXT_COLOR, normalbackground=ACCENT_COLOR,
                                       normalforeground=TEXT_COLOR, headersbackground=BUTTON_COLOR)
    filter_data_inicio_entry.delete(0, tk.END) # Inicia vazio
    filter_data_inicio_entry.pack(side=tk.LEFT, padx=5)
    
    tk.Label(filter_frame, text="Data Fim:", bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=15)
    filter_data_fim_entry = DateEntry(filter_frame, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                                    locale='pt_BR', showweeknumbers=False, selectbackground=BUTTON_COLOR,
                                    selectforeground=TEXT_COLOR, normalbackground=ACCENT_COLOR,
                                    normalforeground=TEXT_COLOR, headersbackground=BUTTON_COLOR)
    filter_data_fim_entry.delete(0, tk.END) # Inicia vazio
    filter_data_fim_entry.pack(side=tk.LEFT, padx=5)

    def aplicar_filtro():
        global LAST_FILTER_START_DATE_BR, LAST_FILTER_END_DATE_BR
        cliente_nome = filter_cliente_var.get()
        cliente_id = CLIENTES_MAP.get(cliente_nome) if cliente_nome != 'Todos' else 'Todos'
        
        data_inicio_dt = None
        data_inicio_br_str = filter_data_inicio_entry.get()
        if data_inicio_br_str:
            try:
                data_inicio_dt = datetime.strptime(data_inicio_br_str, '%d/%m/%Y').date()
                LAST_FILTER_START_DATE_BR = data_inicio_br_str
            except ValueError:
                messagebox.showerror("Erro de Data", "Formato de 'Data Início' inválido. Use DD/MM/AAAA.")
                return
        else:
            LAST_FILTER_START_DATE_BR = ""


        data_fim_dt = None
        data_fim_br_str = filter_data_fim_entry.get()
        if data_fim_br_str:
            try:
                data_fim_dt = datetime.strptime(data_fim_br_str, '%d/%m/%Y').date()
                LAST_FILTER_END_DATE_BR = data_fim_br_str
            except ValueError:
                messagebox.showerror("Erro de Data", "Formato de 'Data Fim' inválido. Use DD/MM/AAAA.")
                return
        else:
            LAST_FILTER_END_DATE_BR = ""

        # Converte para ISO (YYYY-MM-DD) para o DB
        data_inicio_iso = to_iso(datetime.combine(data_inicio_dt, datetime.min.time())) if data_inicio_dt else None
        data_fim_iso = to_iso(datetime.combine(data_fim_dt, datetime.min.time())) if data_fim_dt else None

        atualizar_tabela_atendimentos(tabela_atendimentos, cliente_id=cliente_id, data_inicio_iso=data_inicio_iso, data_fim_iso=data_fim_iso)

    def limpar_filtro_data():
        filter_data_inicio_entry.delete(0, tk.END) 
        filter_data_fim_entry.delete(0, tk.END)   
        filter_cliente_var.set("Todos")
        aplicar_filtro() 

    ttk.Button(filter_frame, text="Aplicar Filtro", command=aplicar_filtro).pack(side=tk.LEFT, padx=10)
    ttk.Button(filter_frame, text="Limpar Filtros", command=limpar_filtro_data).pack(side=tk.LEFT, padx=10)

    # Tabela de Atendimentos
    colunas_atendimentos = ('ID', 'Cliente', 'Data', 'Descrição', 'Valor')
    tabela_atendimentos = ttk.Treeview(visual_frame, columns=colunas_atendimentos, show='headings')
    
    # Adicionar Scrollbar
    scrollbar = ttk.Scrollbar(visual_frame, orient="vertical", command=tabela_atendimentos.yview)
    tabela_atendimentos.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    for col in colunas_atendimentos:
        tabela_atendimentos.heading(col, text=col)
    tabela_atendimentos.column('ID', width=50, stretch=tk.NO, anchor=tk.CENTER)
    tabela_atendimentos.column('Cliente', width=150, anchor=tk.W)
    tabela_atendimentos.column('Data', width=100, anchor=tk.CENTER)
    tabela_atendimentos.column('Valor', width=80, anchor=tk.E)
    tabela_atendimentos.column('Descrição', width=450, anchor=tk.W)
    
    tabela_atendimentos.pack(pady=5, expand=True, fill="both")
    tabela_atendimentos.bind('<<TreeviewSelect>>', preencher_form_atendimento) 


    def atualizar_tabela_atendimentos(tabela, cliente_id=None, data_inicio_iso=None, data_fim_iso=None):
        for i in tabela.get_children():
            tabela.delete(i)

        _, atendimentos = listar_atendimentos(cliente_id, data_inicio_iso, data_fim_iso)
        
        for atd in atendimentos:
            atd_formatado = list(atd[:-1])
            atd_formatado[2] = to_br(atd_formatado[2])
            atd_formatado[4] = f"R$ {atd_formatado[4]:.2f}".replace('.', ',')
            tabela.insert('', 'end', values=atd_formatado)
        
        cliente_nomes = list(CLIENTES_MAP.keys())
        cliente_combo.config(values=cliente_nomes)
        filter_cliente_combo.config(values=["Todos"] + cliente_nomes)
        if cliente_nomes and not cliente_var.get():
             cliente_combo.set(cliente_nomes[0])


    aplicar_filtro() # Aplica o filtro inicial ao carregar
    
    notebook.bind("<<NotebookTabChanged>>", lambda event: aplicar_filtro() if notebook.tab(notebook.select(), "text") == "📞 Registro de Atendimentos" else None)
    
# --------------------------
# --- ABA 2: CLIENTES (SEGUNDA ABA) ---
# --------------------------
frame_clientes = ttk.Frame(notebook)
notebook.add(frame_clientes, text="📝 Clientes Cadastrados")

def criar_aba_clientes(frame):
    
    # Área de Formulário (Cadastro/Edição)
    form_frame = tk.LabelFrame(frame, text=" Cadastro e Edição de Clientes ", padx=10, pady=10, bg=BG_COLOR, fg=TEXT_COLOR)
    form_frame.pack(fill='x', padx=10, pady=5)
    
    campos = ['ID:', 'Nome:', 'CNPJ:', 'Telefone:', 'Endereço:']
    entries = {}
    for i, label_text in enumerate(campos):
        tk.Label(form_frame, text=label_text, width=10, anchor='w', bg=BG_COLOR, fg=TEXT_COLOR).grid(row=i, column=0, padx=5, pady=2, sticky='w')
        entry = tk.Entry(form_frame, width=50, bg=ACCENT_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR)
        entry.grid(row=i, column=1, padx=5, pady=2, sticky='ew')
        entries[label_text.replace(':', '').strip()] = entry
    
    entries['ID'].config(state='disabled', disabledbackground=BG_COLOR, disabledforeground=TEXT_COLOR) 
    
    # --- Funções de Ação (Mantidas as originais) ---
    def limpar_formulario():
        for key, entry in entries.items():
            entry.config(state='normal')
            entry.delete(0, tk.END)
            if key == 'ID':
                entry.config(state='disabled')
        btn_cadastrar.config(text="Cadastrar Novo")


    def preencher_formulario(event):
        try:
            limpar_formulario()
            item_selecionado = tabela_clientes.selection()[0]
            valores = tabela_clientes.item(item_selecionado, 'values')
            
            entries['ID'].config(state='normal')
            entries['ID'].insert(0, valores[0])
            entries['Nome'].insert(0, valores[1])
            entries['CNPJ'].insert(0, valores[2])
            entries['Telefone'].insert(0, valores[3])
            entries['Endereço'].insert(0, valores[4])
            entries['ID'].config(state='disabled')
            
            btn_cadastrar.config(text="Salvar Edição")
        except IndexError:
            limpar_formulario()

    def handle_cadastro_edicao():
        id_cliente = entries['ID'].get()
        nome = entries['Nome'].get()
        cnpj = entries['CNPJ'].get()
        telefone = entries['Telefone'].get()
        endereco = entries['Endereço'].get()
        
        if not nome:
            messagebox.showerror("Erro", "O campo Nome é obrigatório!")
            return
        
        try:
            if id_cliente:
                editar_cliente(id_cliente, nome, cnpj, telefone, endereco)
                messagebox.showinfo("Sucesso", f"Cliente ID {id_cliente} atualizado!")
            else:
                inserir_cliente(nome, cnpj, telefone, endereco)
                messagebox.showinfo("Sucesso", f"Cliente {nome} cadastrado!")
            
            limpar_formulario()
            atualizar_tabela_clientes(tabela_clientes)
            btn_cadastrar.config(text="Cadastrar Novo")
            
        except Exception as e:
            messagebox.showerror("Erro no DB", f"Não foi possível salvar os dados:\n{e}")

    def handle_remover():
        try:
            item_selecionado = tabela_clientes.selection()[0]
            id_cliente = tabela_clientes.item(item_selecionado, 'values')[0]
            nome_cliente = tabela_clientes.item(item_selecionado, 'values')[1]
            
            if messagebox.askyesno("Confirmar Remoção", f"Tem certeza que deseja remover o cliente {nome_cliente} (ID: {id_cliente})?\nIsso apagará todos os seus atendimentos!"):
                remover_cliente(id_cliente)
                messagebox.showinfo("Sucesso", f"Cliente {nome_cliente} removido.")
                limpar_formulario()
                atualizar_tabela_clientes(tabela_clientes)
                notebook.event_generate("<<NotebookTabChanged>>") 
        except IndexError:
            messagebox.showwarning("Atenção", "Selecione um cliente para remover.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível remover o cliente:\n{e}")

    def handle_importar_excel():
        filepath = filedialog.askopenfilename(
            title="Selecione o arquivo Excel de Clientes",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if filepath:
            top = tk.Toplevel(root)
            top.title("Importando...")
            tk.Label(top, text="Importação em andamento, por favor aguarde...", bg=BG_COLOR, fg=TEXT_COLOR).pack(padx=20, pady=20)
            root.update_idletasks() 
            
            importados, mensagem = importar_clientes_excel(filepath)
            top.destroy()
            
            if importados > 0:
                messagebox.showinfo("Importação Concluída", mensagem)
                atualizar_tabela_clientes(tabela_clientes)
                notebook.event_generate("<<NotebookTabChanged>>")
            else:
                messagebox.showerror("Erro de Importação", mensagem)


    # Botões de Ação
    btn_frame = tk.Frame(form_frame, bg=BG_COLOR)
    btn_frame.grid(row=len(campos), column=0, columnspan=2, pady=10)
    
    btn_cadastrar = ttk.Button(btn_frame, text="Cadastrar Novo", command=handle_cadastro_edicao)
    btn_cadastrar.pack(side=tk.LEFT, padx=5)
    
    ttk.Button(btn_frame, text="Limpar", command=limpar_formulario).pack(side=tk.LEFT, padx=5)
    ttk.Button(btn_frame, text="Remover Selecionado", command=handle_remover, style='Danger.TButton').pack(side=tk.LEFT, padx=5)
    
    ttk.Button(btn_frame, text="📥 Importar de Excel", command=handle_importar_excel, style='Accent.TButton').pack(side=tk.LEFT, padx=20)


    # Tabela de Visualização
    colunas_clientes, _ = listar_clientes() 
    tabela_clientes = ttk.Treeview(frame, columns=colunas_clientes, show='headings')
    
    # Adicionar Scrollbar
    scrollbar_clientes = ttk.Scrollbar(frame, orient="vertical", command=tabela_clientes.yview)
    tabela_clientes.configure(yscrollcommand=scrollbar_clientes.set)
    scrollbar_clientes.pack(side="right", fill="y")


    for i, col in enumerate(colunas_clientes):
        tabela_clientes.heading(col, text=col)
        tabela_clientes.column(col, width=100 if i > 0 else 50, anchor=tk.W)
        if col == 'ID':
            tabela_clientes.column(col, width=50, stretch=tk.NO, anchor=tk.CENTER)

    tabela_clientes.pack(pady=10, padx=10, expand=True, fill="both")
    tabela_clientes.bind('<<TreeviewSelect>>', preencher_formulario)
    
    def atualizar_tabela_clientes(tabela):
        global CLIENTES_MAP, CLIENTES_ID_MAP
        for i in tabela.get_children():
            tabela.delete(i)
        
        _, clientes = listar_clientes() 
        CLIENTES_MAP = {} 
        CLIENTES_ID_MAP = {}
        
        for cliente in clientes:
            tabela.insert('', 'end', values=cliente)
            CLIENTES_MAP[cliente[1]] = cliente[0] 
            CLIENTES_ID_MAP[cliente[0]] = cliente[1]

    atualizar_tabela_clientes(tabela_clientes)
    
# --- Ordem de criação ---
criar_aba_atendimentos(frame_atendimentos)
criar_aba_clientes(frame_clientes)

# --- Iniciar a Aplicação ---
if __name__ == '__main__':
    _, clientes_list = listar_clientes()
    CLIENTES_MAP = {c[1]: c[0] for c in clientes_list}
    CLIENTES_ID_MAP = {c[0]: c[1] for c in clientes_list}
    root.mainloop()